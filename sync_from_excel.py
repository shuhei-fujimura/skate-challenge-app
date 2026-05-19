from __future__ import annotations

import json
import urllib.parse
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
APP_DIR = Path(__file__).resolve().parent
XLSX = ROOT / "outputs" / "skate_tracker" / "スケートボード練習記録テンプレート.xlsx"
DATA_DIR = APP_DIR / "data"
TRICKS_JSON = DATA_DIR / "tricks.json"
META_JSON = DATA_DIR / "sync_meta.json"


def youtube_url(name: str) -> str:
    query = urllib.parse.quote(f"スケボー {name} やり方")
    return f"https://www.youtube.com/results?search_query={query}"


def row_is_trick(name: object, kind: object) -> bool:
    if not name or not kind:
        return False
    text = str(name)
    if text == "使い方メモ":
        return False
    skip_prefixes = ("毎日版", "毎週版", "動画列", "印刷用シート", "痛いところ", "新しい技")
    return not any(text.startswith(prefix) for prefix in skip_prefixes)


def main() -> None:
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["技リスト"]
    tricks: list[dict[str, object]] = []

    for row in range(4, ws.max_row + 1):
        name = ws[f"B{row}"].value
        kind = ws[f"C{row}"].value
        if not row_is_trick(name, kind):
            continue
        trick_name = str(name).strip()
        tricks.append(
            {
                "id": f"trick-{len(tricks) + 1:03d}",
                "no": len(tricks) + 1,
                "name": trick_name,
                "kind": str(kind).strip(),
                "level": str(ws[f"D{row}"].value or "").strip(),
                "target": str(ws[f"E{row}"].value or "").strip(),
                "memo": str(ws[f"F{row}"].value or "").strip(),
                "video": youtube_url(trick_name),
            }
        )

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    TRICKS_JSON.write_text(json.dumps(tricks, ensure_ascii=False, indent=2), encoding="utf-8")
    META_JSON.write_text(
        json.dumps(
            {
                "source": XLSX.name,
                "sheet": "技リスト",
                "count": len(tricks),
                "synced_at": datetime.now().isoformat(timespec="seconds"),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"synced {len(tricks)} tricks")


if __name__ == "__main__":
    main()
